import openpyxl

def load_active_sheet(res=None):
    res = res if res is not None else 'sheet.xlsx'
    document_workbook = openpyxl.load_workbook(res)
    sheet = document_workbook.get_sheet_by_name(
        document_workbook.get_sheet_names()[0])
    return sheet


def get_rows(start_index, stop_index):
    sheet = load_active_sheet()

    all_data = []

    for row in range(start_index, stop_index):
        row_data = []

        for col in range(1, 3):
            row_data.append(str(sheet.cell(row=row, column=col).value))
        row_data.append(0)
        
        #all_data.append(tuple(row_data))
        all_data.append(row_data)

    return all_data    



def load_db_to_sheet():
    from audit import db_ops

    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.active

    all_data = db_ops.ret_all(db_ops.Record)

    for data in all_data:
        row = [data.isbn, data.name, data.units]
        new_worksheet.append(row)

    new_workbook.save('audit_output_sheet.xlsx')

# 7 - 5875